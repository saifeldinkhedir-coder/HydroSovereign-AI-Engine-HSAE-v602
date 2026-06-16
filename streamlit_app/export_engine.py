"""
export_engine.py — HSAE v10.0 Multi-Format Export Engine
=========================================================
Export HSAE results in professional scientific formats:
  1. NetCDF-4  — time series (CF-1.8 compliant)
  2. Excel     — multi-sheet with charts (openpyxl)
  3. GeoJSON   — basin geometries + ATDI attributes
  4. Shapefile — compatible GIS output (shapely + pyproj)
  5. CSV       — universal fallback
  6. HTML      — standalone report (already in pdf_exporter)

Author: Seifeldin M.G. Alkhedir · ORCID: 0000-0003-0821-2991
"""
from __future__ import annotations
import io
import os
import json
import csv
import datetime
from typing import Dict, List, Optional, Any

# Optional imports
try:
    import numpy as np
    _NP = True
except ImportError:
    _NP = False

try:
    import pandas as pd
    _PD = True
except ImportError:
    _PD = False

try:
    import netCDF4 as nc
    _NC = True
except ImportError:
    _NC = False

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ══════════════════════════════════════════════════════════════════════════════
# 1. NetCDF-4 Export (CF-1.8 compliant)
# ══════════════════════════════════════════════════════════════════════════════
def export_netcdf(
    df,
    basin: Dict,
    variables: Optional[List[str]] = None,
    output_path: Optional[str] = None,
) -> bytes:
    """
    Export time series DataFrame to CF-1.8 compliant NetCDF-4.

    Parameters
    ----------
    df           : pandas DataFrame with Date column
    basin        : basin config dict
    variables    : columns to export (None = all numeric)
    output_path  : save to file if given; else return bytes

    Returns bytes if output_path is None.
    """
    if not _NC or not _PD or not _NP:
        # Fallback: CSV as bytes
        return df.to_csv(index=False).encode("utf-8")

    basin_id = basin.get("id", "UNKNOWN")
    buf_path  = output_path or str(__import__("pathlib").Path(__import__("tempfile").gettempdir()) / f"hsae_{basin_id}_{datetime.datetime.utcnow().strftime('%Y%m%d%H%M%S')}.nc")

    df_clean = df.copy()
    if "Date" in df_clean.columns:
        df_clean["Date"] = pd.to_datetime(df_clean["Date"])
        df_clean = df_clean.sort_values("Date")

    if variables is None:
        variables = [c for c in df_clean.columns
                     if c != "Date" and pd.api.types.is_numeric_dtype(df_clean[c])]

    with nc.Dataset(buf_path, "w", format="NETCDF4") as ds:
        # Global attributes (CF-1.8)
        ds.Conventions      = "CF-1.8"
        ds.title            = f"HSAE v10.0 — {basin.get('_name', basin_id)}"
        ds.institution      = "University of Khartoum"
        ds.source           = "HydroSovereign AI Engine v10.0"
        ds.history          = f"Created {datetime.datetime.utcnow().isoformat()} by HSAE v10.0"
        ds.references       = "Alkhedir (2026) doi:PENDING_RELEASE"
        ds.basin_id         = basin_id
        ds.basin_name       = str(basin.get("_name", basin_id))
        ds.basin_river      = str(basin.get("river", ""))
        ds.basin_cap_BCM    = float(basin.get("cap", 0))
        ds.creator_name     = "Seifeldin M.G. Alkhedir"
        ds.creator_orcid    = "0000-0003-0821-2991"
        ds.geospatial_lat   = float(basin.get("lat", 0))
        ds.geospatial_lon   = float(basin.get("lon", 0))

        # Time dimension
        n_time = len(df_clean)
        ds.createDimension("time", n_time)

        if "Date" in df_clean.columns:
            t_var = ds.createVariable("time", "f8", ("time",))
            t_var.units     = "days since 2000-01-01 00:00:00"
            t_var.calendar  = "standard"
            t_var.axis      = "T"
            t_var.long_name = "time"
            origin = datetime.datetime(2000, 1, 1)
            t_var[:] = np.array([
                (d.to_pydatetime() - origin).total_seconds() / 86400
                for d in pd.to_datetime(df_clean["Date"])
            ])

        # Variable attributes map
        var_attrs = {
            "Inflow_BCM":     {"long_name": "Reservoir inflow",           "units": "BCM"},
            "Outflow_BCM":    {"long_name": "Reservoir outflow",          "units": "BCM"},
            "Volume_BCM":     {"long_name": "Reservoir storage volume",   "units": "BCM"},
            "GPM_Rain_mm":    {"long_name": "GPM IMERG rainfall",         "units": "mm/day"},
            "S1_VV_dB":       {"long_name": "Sentinel-1 VV backscatter",  "units": "dB"},
            "S2_NDWI":        {"long_name": "Sentinel-2 NDWI",            "units": "1"},
            "Effective_Area": {"long_name": "Water surface area (fused)", "units": "km2"},
            "Q_obs":          {"long_name": "Observed discharge",         "units": "m3 s-1"},
            "Q_sim":          {"long_name": "Simulated discharge",        "units": "m3 s-1"},
            "atdi":           {"long_name": "Alkhedir Transboundary Dependency Index", "units": "1"},
        }

        for vname in variables:
            if vname not in df_clean.columns:
                continue
            data = df_clean[vname].fillna(-9999.0).to_numpy()
            v    = ds.createVariable(vname, "f4", ("time",),
                                     fill_value=-9999.0, zlib=True, complevel=4)
            attrs = var_attrs.get(vname, {})
            v.long_name    = attrs.get("long_name", vname)
            v.units        = attrs.get("units", "1")
            v.missing_value= -9999.0
            v[:]           = data

    if output_path:
        return b""  # saved to file

    with open(buf_path, "rb") as f:
        data_bytes = f.read()
    os.unlink(buf_path)
    return data_bytes


# ══════════════════════════════════════════════════════════════════════════════
# 2. Excel Multi-Sheet Export (openpyxl)
# ══════════════════════════════════════════════════════════════════════════════
# Colour theme
_COL_DARK   = "0D1117"
_COL_GREEN  = "10B981"
_COL_BLUE   = "3B82F6"
_COL_ORANGE = "F97316"
_COL_TEXT   = "E6EDF3"
_COL_SUBHDR = "161B22"


def _xl_header_style(ws, row: int, cols: int, title: str):
    """Apply HSAE dark header style to a row."""
    if not _OPENPYXL:
        return
    ws.row_dimensions[row].height = 25
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = PatternFill("solid", fgColor=_COL_DARK)
        cell.font      = Font(color=_COL_GREEN, bold=True, name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(bottom=Side(style="thin", color=_COL_GREEN))
    ws.cell(row=row, column=1).value = title


def export_excel(
    df,
    basin: Dict,
    results: Optional[Dict] = None,
    dossier: Optional[Dict] = None,
) -> bytes:
    """
    Export HSAE results to a professional multi-sheet Excel workbook.

    Sheets:
      1. Summary        — basin info + key metrics
      2. Time Series    — full DataFrame
      3. Legal          — treaty violations + ICJ evidence
      4. Benchmark      — NSE/KGE vs published values
      5. Parameters     — HBV parameters (if available)
    """
    if not _OPENPYXL or not _PD:
        # Fallback: simple CSV bytes
        return df.to_csv(index=False).encode("utf-8") if _PD else b""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    results = results or {}
    basin_id = basin.get("id", basin.get("_v9_id", "UNKNOWN"))

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False

    # Logo row
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"🌊  HydroSovereign AI Engine v10.0 — {basin.get('_name', basin_id)}"
    ws1["A1"].font      = Font(bold=True, size=16, color=_COL_GREEN, name="Calibri")
    ws1["A1"].fill      = PatternFill("solid", fgColor=_COL_DARK)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 35

    ws1["A2"] = f"Generated: {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws1["A2"].font = Font(color="8B949E", italic=True, name="Calibri")

    # Basin info
    ws1["A4"] = "Basin Information"
    ws1["A4"].font = Font(bold=True, size=13, color=_COL_BLUE, name="Calibri")

    basin_info = [
        ("Basin ID",    basin_id),
        ("River",       basin.get("river", "—")),
        ("Dam",         basin.get("dam",   "—")),
        ("Country",     ", ".join(basin.get("country", [])[:3])),
        ("Continent",   basin.get("continent", "—")),
        ("Capacity",    f"{basin.get('cap', '—')} BCM"),
        ("Head",        f"{basin.get('head', '—')} m"),
        ("Treaty",      str(basin.get("treaty", "—"))[:40]),
    ]
    for i, (lbl, val) in enumerate(basin_info, start=5):
        ws1[f"A{i}"] = lbl
        ws1[f"B{i}"] = val
        ws1[f"A{i}"].font = Font(color=_COL_TEXT, bold=True, name="Calibri")
        ws1[f"B{i}"].font = Font(color="34D399", name="Calibri")
        ws1[f"A{i}"].fill = PatternFill("solid", fgColor=_COL_SUBHDR)
        ws1[f"B{i}"].fill = PatternFill("solid", fgColor=_COL_SUBHDR)

    # Key metrics
    ws1["D4"] = "Key Indices"
    ws1["D4"].font = Font(bold=True, size=13, color=_COL_ORANGE, name="Calibri")

    metrics = [
        ("ATDI",   results.get("atdi",  results.get("legal", {}).get("atdi", "—"))),
        ("AHIFD",  f"{results.get('ahifd_pct', '—')}%"),
        ("ATCI",   f"{results.get('atci', '—')}/100"),
        ("NSE",    results.get("nse",   "—")),
        ("KGE",    results.get("kge",   "—")),
        ("PBIAS",  f"{results.get('pbias', '—')}%"),
    ]
    for i, (lbl, val) in enumerate(metrics, start=5):
        ws1[f"D{i}"] = lbl
        ws1[f"E{i}"] = val
        ws1[f"D{i}"].font = Font(color=_COL_TEXT, bold=True, name="Calibri")
        ws1[f"E{i}"].font = Font(color="FCD34D", name="Calibri")
        ws1[f"D{i}"].fill = PatternFill("solid", fgColor=_COL_SUBHDR)
        ws1[f"E{i}"].fill = PatternFill("solid", fgColor=_COL_SUBHDR)

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 20

    # ── Sheet 2: Time Series ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Time Series")
    ws2.sheet_view.showGridLines = False

    if _PD and df is not None and not df.empty:
        cols = df.columns.tolist()
        _xl_header_style(ws2, 1, len(cols), "")
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws2.cell(row=1, column=c_idx)
            cell.value     = col_name
            cell.font      = Font(color=_COL_GREEN, bold=True, name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=_COL_DARK)
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx)
                if hasattr(val, "isoformat"):
                    cell.value = val.isoformat()[:10]
                elif isinstance(val, float):
                    cell.value = round(val, 4) if not (
                        val != val or abs(val) == float("inf")) else None
                else:
                    cell.value = val
                cell.font = Font(name="Calibri", size=9)
                if r_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="0F172A")

        for c_idx in range(1, len(cols)+1):
            ws2.column_dimensions[get_column_letter(c_idx)].width = 14

    # ── Sheet 3: Legal ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Legal Assessment")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "UN 1997 Watercourses Convention — Compliance Assessment"
    ws3["A1"].font = Font(bold=True, size=13, color=_COL_BLUE, name="Calibri")
    ws3["A1"].fill = PatternFill("solid", fgColor=_COL_DARK)
    ws3.merge_cells("A1:F1")

    if dossier:
        headers = ["Article", "Title", "Gravity", "Obligation", "Status"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws3.cell(row=2, column=c_idx)
            cell.value = h
            cell.font  = Font(color=_COL_GREEN, bold=True, name="Calibri")
            cell.fill  = PatternFill("solid", fgColor=_COL_SUBHDR)

        for r_idx, v in enumerate(dossier.get("articles_triggered", [])[:20], 3):
            ws3.cell(row=r_idx, column=1).value = v.get("article", "")
            ws3.cell(row=r_idx, column=2).value = v.get("title", "")[:50]
            ws3.cell(row=r_idx, column=3).value = v.get("gravity", "")
            ws3.cell(row=r_idx, column=4).value = v.get("text", "")[:80]
            ws3.cell(row=r_idx, column=5).value = "TRIGGERED"
            grav = v.get("gravity", "")
            color = "F85149" if grav == "HIGH" else "F97316" if grav == "MEDIUM" else "10B981"
            ws3.cell(row=r_idx, column=5).font = Font(color=color, bold=True)

    for c_idx, width in enumerate([12, 35, 12, 50, 12], 1):
        ws3.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Sheet 4: Benchmark ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Benchmark")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "HSAE v10.0 Benchmark Comparison (Moriasi et al. 2007)"
    ws4["A1"].font = Font(bold=True, size=13, color=_COL_ORANGE, name="Calibri")
    ws4["A1"].fill = PatternFill("solid", fgColor=_COL_DARK)
    ws4.merge_cells("A1:G1")

    try:
        from benchmark_comparison import LITERATURE_BENCHMARKS, moriasi_rating
        headers_b = ["Basin", "Metric", "HSAE", "Best Published", "DOI", "Rating"]
        for c_idx, h in enumerate(headers_b, 1):
            ws4.cell(row=2, column=c_idx).value = h
            ws4.cell(row=2, column=c_idx).font = Font(color=_COL_GREEN, bold=True, name="Calibri")

        row_b = 3
        for b_id, data in list(LITERATURE_BENCHMARKS.items())[:10]:
            for metric, entries in data.items():
                if isinstance(entries, dict):
                    for key, rec in entries.items():
                        if isinstance(rec, dict) and rec.get("value") is not None:
                            ws4.cell(row=row_b, column=1).value = b_id
                            ws4.cell(row=row_b, column=2).value = metric
                            ws4.cell(row=row_b, column=3).value = results.get(metric.lower(), "—")
                            ws4.cell(row=row_b, column=4).value = rec["value"]
                            ws4.cell(row=row_b, column=5).value = rec.get("doi", "")
                            ws4.cell(row=row_b, column=6).value = moriasi_rating(
                                float(rec["value"]) if isinstance(rec["value"], (int,float)) else 0.5,
                                0.7, 10)
                            row_b += 1
    except Exception:
        ws4["A3"] = "Benchmark data: see benchmark_comparison.py"

    for c_idx, width in enumerate([20, 10, 10, 15, 30, 15], 1):
        ws4.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Finalize ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# 3. GeoJSON Export
# ══════════════════════════════════════════════════════════════════════════════
def export_geojson(
    basins: Dict,
    atdi_values: Optional[Dict[str, float]] = None,
) -> bytes:
    """Export basin geometries + ATDI attributes as GeoJSON FeatureCollection."""
    features = []
    for name, cfg in basins.items():
        lat = cfg.get("lat", 0)
        lon = cfg.get("lon", 0)
        atdi = (atdi_values or {}).get(
            cfg.get("id", name), cfg.get("tdi", 0.5))

        feature = {
            "type": "Feature",
            "geometry": {
                "type": "Point",
                "coordinates": [lon, lat]
            },
            "properties": {
                "name":      name,
                "id":        cfg.get("id", ""),
                "river":     cfg.get("river", ""),
                "dam":       cfg.get("dam", ""),
                "continent": cfg.get("continent", ""),
                "country":   cfg.get("country", [""]),
                "capacity_BCM": cfg.get("cap", 0),
                "head_m":    cfg.get("head", 0),
                "treaty":    str(cfg.get("treaty", "")),
                "atdi":      round(float(atdi), 3) if atdi else None,
                "risk_level": ("HIGH" if float(atdi or 0) > 0.70 else
                               "MODERATE" if float(atdi or 0) > 0.45 else "LOW"),
            }
        }
        features.append(feature)

    geojson = {
        "type": "FeatureCollection",
        "name": "HSAE_v10_Basins",
        "crs":  {"type": "name",
                 "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}},
        "features": features
    }
    return json.dumps(geojson, indent=2, ensure_ascii=False).encode("utf-8")


# ══════════════════════════════════════════════════════════════════════════════
# 4. Universal CSV Export (fallback)
# ══════════════════════════════════════════════════════════════════════════════
def export_csv(df, encoding: str = "utf-8-sig") -> bytes:
    """Export DataFrame to CSV bytes (utf-8-sig for Excel compatibility)."""
    if _PD and df is not None:
        return df.to_csv(index=False, encoding=encoding).encode(encoding)
    return b""


# ══════════════════════════════════════════════════════════════════════════════
# 5. Streamlit export panel
# ══════════════════════════════════════════════════════════════════════════════
def render_export_panel(
    basin: Dict,
    df=None,
    results: Optional[Dict] = None,
    dossier: Optional[Dict] = None,
):
    """Complete Streamlit export panel with all formats."""
    try:
        import streamlit as st
    except ImportError:
        return

    st.subheader("⬇️ Export Results")
    basin_id = basin.get("id", basin.get("_v9_id", "basin"))
    ts       = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M")

    c1, c2, c3, c4 = st.columns(4)

    # CSV — always works
    with c1:
        if df is not None:
            st.download_button(
                "📄 CSV",
                data=export_csv(df),
                file_name=f"HSAE_{basin_id}_{ts}.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # Excel — needs openpyxl
    with c2:
        if _OPENPYXL and df is not None:
            try:
                xlsx_bytes = export_excel(df, basin, results, dossier)
                st.download_button(
                    "📊 Excel",
                    data=xlsx_bytes,
                    file_name=f"HSAE_{basin_id}_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.caption(f"Excel: {e}")
        else:
            st.caption("Excel: `pip install openpyxl`")

    # NetCDF
    with c3:
        if _NC and df is not None:
            try:
                nc_bytes = export_netcdf(df, basin)
                st.download_button(
                    "🌐 NetCDF",
                    data=nc_bytes,
                    file_name=f"HSAE_{basin_id}_{ts}.nc",
                    mime="application/octet-stream",
                    use_container_width=True,
                )
            except Exception as e:
                st.caption(f"NetCDF: {e}")
        else:
            st.caption("NetCDF: `pip install netCDF4`")

    # GeoJSON
    with c4:
        try:
            from basins_bridge import UNIFIED_BASINS
            gj_bytes = export_geojson(UNIFIED_BASINS)
            st.download_button(
                "🗺️ GeoJSON",
                data=gj_bytes,
                file_name=f"HSAE_basins_{ts}.geojson",
                mime="application/geo+json",
                use_container_width=True,
            )
        except Exception as e:
            st.caption(f"GeoJSON: {e}")

    # HTML Report
    st.markdown("---")
    try:
        from pdf_exporter import generate_pdf_report, render_pdf_export_panel
        render_pdf_export_panel(basin, results or {}, dossier)
    except Exception:
        pass
